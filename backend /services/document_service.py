"""
Генерация Excel-документов.
"""
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from sqlmodel import Session, select
import num2words

from backend.db import get_session
from backend.models.order import Order
from backend.models.client import Client
from backend.models.vehicle import Vehicle
from backend.config import load_json, DATA_DIR


class DocumentService:

    def __init__(self):
        self.output_dir = DATA_DIR / "documents"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_amount_in_words(self, amount: float) -> str:
        """Сумма прописью."""
        try:
            rubles = int(amount)
            kopecks = round((amount - rubles) * 100)

            if kopecks == 0:
                kop_str = "00 копеек"
            elif 11 <= kopecks <= 19:
                kop_str = f"{kopecks} копеек"
            elif kopecks % 10 == 1:
                kop_str = f"{kopecks} копейка"
            elif 2 <= kopecks % 10 <= 4:
                kop_str = f"{kopecks} копейки"
            else:
                kop_str = f"{kopecks} копеек"

            rub_words = num2words.num2words(rubles, lang='ru')

            if 11 <= rubles % 100 <= 19:
                rub_word = "рублей"
            elif rubles % 10 == 1:
                rub_word = "рубль"
            elif 2 <= rubles % 10 <= 4:
                rub_word = "рубля"
            else:
                rub_word = "рублей"

            return f"{rub_words.capitalize()} {rub_word} {kop_str}"
        except:
            return f"{amount:.2f} руб."

    def generate_order_excel(self, order_id: str) -> str:
        """Генерирует Excel заказ-наряда. Возвращает путь к файлу."""

        with get_session() as session:
            # Загружаем заказ
            order = session.exec(
                select(Order).where(Order.order_id == order_id)
            ).first()

            if not order:
                raise ValueError("Заказ не найден")

            # Загружаем клиента
            client = session.exec(
                select(Client).where(Client.id == order.client_id)
            ).first()

            # Загружаем авто
            vehicle = session.exec(
                select(Vehicle).where(Vehicle.id == order.vehicle_id)
            ).first()

            # Реквизиты компании
            contractor = load_json("contractor.json")

            # Данные заказа
            works = order.works
            materials = order.materials
            total = order.total_amount

            # Имя файла
            plate = vehicle.plate if vehicle else "____"
            date_clean = order.date.replace('.', '-')
            filename = f"ZN_{order.zn_number}_{plate}_{date_clean}.xlsx"
            filepath = self.output_dir / filename

            # ===== Excel =====
            wb = Workbook()
            ws = wb.active
            ws.title = "Заказ-наряд"

            # Стили
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            row = 1

            # === ШАПКА ===
            company = contractor.get("company", "ИП")
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f"ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ    {company.replace('ИП ', '')}"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 1

            inn = contractor.get("inn", "")
            ogrnip = contractor.get("ogrnip", "")
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f"ИНН: {inn} ОГРНИП: {ogrnip}"
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 1

            address = contractor.get("address", "")
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = address
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 1

            email = contractor.get("email", "")
            phone = contractor.get("phone", "")
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f"{email} {phone}"
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2

            # === ЗАКАЗ-НАРЯД ===
            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = f"ЗАКАЗ – НАРЯД № {order.zn_number}"
            ws[f'B{row}'].font = Font(bold=True, size=14)
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = f"Дата и время приема заказа: {order.date} г."
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = f"Дата и время окончания работ: {order.date} г."
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1

            # === ЗАКАЗЧИК ===
            client_name = client.full_name if client else "—"
            client_addr = client.address if client else "—"

            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = "Заказчик"
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = client_name
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = f"Адрес: {client_addr}"
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1

            # === АВТОМОБИЛЬ ===
            v_brand = vehicle.brand if vehicle else ""
            v_model = vehicle.model if vehicle else ""
            v_plate = vehicle.plate if vehicle else "__________"
            v_year = str(vehicle.year) if vehicle and vehicle.year else ""

            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'] = f"Марка, модель: {v_brand} {v_model}"
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            ws[f'E{row}'] = "Двигатель №"
            ws[f'E{row}'].alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'] = f"Гос. рег. номер: {v_plate}"
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            ws[f'E{row}'] = "Шасси №"
            ws[f'E{row}'].alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'] = f"Год выпуска: {v_year}"
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            ws[f'E{row}'] = "Кузов №"
            ws[f'E{row}'].alignment = Alignment(horizontal='center')
            row += 1

            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = f"VIN: {vehicle.vin if vehicle else ''}"
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            row += 1

            # === РАБОТЫ ===
            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = f"Выполненные работы по заказ–наряду № {order.zn_number}"
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1

            # Заголовки таблицы работ
            work_headers = ["№", "Наименование работ", "Норма времени", "Кол-во", "Стоимость (руб.)", "Сумма (руб.)"]
            for col, h in enumerate(work_headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[row].height = 30
            work_start_row = row
            row += 1

            # Данные работ
            work_first_data = row
            for i, w in enumerate(works, 1):
                ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=2, value=w.get('name', '')).alignment = Alignment(horizontal='left')
                ws.cell(row=row, column=3, value=w.get('norm_hours', 0)).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=4, value=w.get('quantity', 1)).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=w.get('rate_rub', 0)).alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=6, value=w.get('sum_rub', 0)).alignment = Alignment(horizontal='center')
                row += 1
            work_last_data = row - 1

            # Итого работы
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'] = "Итого работы (руб.)"
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            ws[f'F{row}'] = f"=SUM(F{work_first_data}:F{work_last_data})"
            ws[f'F{row}'].font = Font(bold=True)
            ws[f'F{row}'].alignment = Alignment(horizontal='center')
            work_total_row = row
            row += 2

            # === ЗАПЧАСТИ (если есть) ===
            if materials:
                ws.merge_cells(f'B{row}:F{row}')
                ws[f'B{row}'] = f"Расходная накладная по заказ–наряду № {order.zn_number}"
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'B{row}'].alignment = Alignment(horizontal='center')
                row += 1

                part_headers = ["№", "Наименование", "Ед.", "Кол-во", "Цена (руб.)", "Сумма (руб.)"]
                for col, h in enumerate(part_headers, 1):
                    cell = ws.cell(row=row, column=col, value=h)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1

                part_first_data = row
                for i, m in enumerate(materials, 1):
                    ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=2, value=m.get('name', '')).alignment = Alignment(horizontal='left')
                    ws.cell(row=row, column=3, value=m.get('unit', 'шт.')).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=4, value=m.get('quantity', 1)).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=5, value=m.get('cost_rub', 0)).alignment = Alignment(horizontal='center')
                    ws.cell(row=row, column=6, value=f"=D{row}*E{row}").alignment = Alignment(horizontal='center')
                    row += 1
                part_last_data = row - 1

                ws.merge_cells(f'B{row}:E{row}')
                ws[f'B{row}'] = "Итого запчасти (руб.)"
                ws[f'B{row}'].font = Font(bold=True)
                ws[f'B{row}'].alignment = Alignment(horizontal='left')
                ws[f'F{row}'] = f"=SUM(F{part_first_data}:F{part_last_data})"
                ws[f'F{row}'].font = Font(bold=True)
                ws[f'F{row}'].alignment = Alignment(horizontal='center')
                part_total_row = row
                row += 2
            else:
                part_total_row = 0

            # === ИТОГО ===
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'] = "Всего к оплате (руб.)"
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            if part_total_row > 0:
                ws[f'F{row}'] = f"=F{work_total_row}+F{part_total_row}"
            else:
                ws[f'F{row}'] = f"=F{work_total_row}"
            ws[f'F{row}'].font = Font(bold=True)
            ws[f'F{row}'].alignment = Alignment(horizontal='center')
            row += 2

            # Сумма прописью
            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = f"Всего по заказ-наряду: {self._get_amount_in_words(total)}"
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
            row += 2

            # === ПОДПИСИ ===
            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = "Заказчик________________ МП Исполнитель_______________ МП"
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 2

            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'] = "Работы выполнены с использованием запасных частей заказчика"
            ws[f'B{row}'].alignment = Alignment(horizontal='center')

            # === ШИРИНА СТОЛБЦОВ ===
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 14

            # Сохраняем
            wb.save(filepath)

            return filename