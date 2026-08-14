# backend/engines/full_document_xlsx.py
"""Генератор полного пакета документов: Заказ-наряд + Счёт + УПД"""
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import num2words


class FullDocumentGenerator:
    """Генерирует Excel с тремя листами в старом стиле"""
    
    def __init__(self, project_root: Path = None):
        if project_root is None:
            project_root = Path(__file__).parent.parent.parent
        
        self.output_dir = project_root / "data" / "documents"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Загружаем реквизиты
        contractor_path = project_root / "config" / "contractor.json"
        with open(contractor_path, encoding="utf-8") as f:
            self.contractor = json.load(f)
        
        # Загружаем клиентов (для адресов)
        client_path = project_root / "config" / "client.json"
        if client_path.exists():
            with open(client_path, encoding="utf-8") as f:
                self.clients = json.load(f).get("clients", {})
        else:
            self.clients = {}
        
        # Стили
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        self.bold_font = Font(bold=True, size=11)
        self.normal_font = Font(size=11)
        self.title_font = Font(bold=True, size=12)
        self.center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.right = Alignment(horizontal='right', vertical='center')
    
    def _amount_in_words(self, amount: float) -> str:
        """Сумма прописью"""
        rubles = int(amount)
        kopecks = int(round((amount - rubles) * 100))
        words = num2words.num2words(rubles, lang='ru')
        last = rubles % 10
        last2 = rubles % 100
        if 11 <= last2 <= 19:
            rub_word = "рублей"
        elif last == 1:
            rub_word = "рубль"
        elif 2 <= last <= 4:
            rub_word = "рубля"
        else:
            rub_word = "рублей"
        return f"{words.capitalize()} {rub_word} {kopecks:02d} копеек."
    
    def _format_plate(self, plate: str) -> str:
        """Форматирует госномер с пробелом"""
        if not plate:
            return ""
        clean = plate.upper().replace(' ', '')
        if len(clean) >= 6:
            return f"{clean[:-4]} {clean[-4:]}"
        return clean
    
    def _apply_border(self, ws, min_row, max_row, min_col, max_col):
        """Применяет границы к диапазону"""
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).border = self.thin_border
    
    def _write_company_header(self, ws, current_row):
        """Шапка с реквизитами ИП (как в старом заказ-наряде)"""
        company = self.contractor.get('company', '').replace('ИП ', '')
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ    {company}"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].alignment = self.center
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"ИНН: {self.contractor.get('inn', '')} ОГРНИП: {self.contractor.get('ogrnip', '')}"
        ws[f'A{current_row}'].alignment = self.center
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = self.contractor.get('address', '')
        ws[f'A{current_row}'].alignment = self.center
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"{self.contractor.get('email', '')} {self.contractor.get('phone', '')}"
        ws[f'A{current_row}'].alignment = self.center
        current_row += 1
        
        return current_row
    
    def _write_work_table(self, ws, current_row, works: list):
        """Таблица работ (старый формат)"""
        current_row += 1
        headers = ["№", "Наименование работ", "Норма времени", "Кол-во", "Стоимость (руб.)", "Сумма (руб.)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.header_fill
            cell.alignment = self.center
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        first_work_row = current_row
        for i, work in enumerate(works, 1):
            ws.cell(row=current_row, column=1, value=i).alignment = self.center
            ws.cell(row=current_row, column=2, value=work['name']).alignment = self.left
            ws.cell(row=current_row, column=3, value=work['norm_hours']).alignment = self.center
            ws.cell(row=current_row, column=4, value=work.get('quantity', 1)).alignment = self.center
            ws.cell(row=current_row, column=5, value=work['rate_rub']).alignment = self.center
            ws.cell(row=current_row, column=6, value=f"=C{current_row}*D{current_row}*E{current_row}").alignment = self.center
            current_row += 1
        last_work_row = current_row - 1
        
        ws.merge_cells(f"B{current_row}:E{current_row}")
        ws[f"B{current_row}"] = "Итого работы (руб.)"
        ws[f"B{current_row}"].font = self.bold_font
        ws[f"B{current_row}"].alignment = self.left
        ws[f"F{current_row}"] = f"=SUM(F{first_work_row}:F{last_work_row})"
        ws[f"F{current_row}"].font = self.bold_font
        ws[f"F{current_row}"].alignment = self.center
        
        self._apply_border(ws, first_work_row - 1, current_row, 1, 6)
        return current_row
    
    def _write_material_table(self, ws, current_row, materials: list):
        """Таблица материалов (старый формат)"""
        if not materials:
            return current_row
        
        current_row += 2
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws[f"B{current_row}"] = "Расходная накладная"
        ws[f"B{current_row}"].font = self.bold_font
        ws[f"B{current_row}"].alignment = self.center
        current_row += 1
        
        headers = ["№", "Наименование", "Ед.", "Кол-во", "Цена (руб.)", "Сумма (руб.)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = self.header_fill
            cell.alignment = self.center
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        first_mat_row = current_row
        for i, material in enumerate(materials, 1):
            ws.cell(row=current_row, column=1, value=i).alignment = self.center
            ws.cell(row=current_row, column=2, value=material['name']).alignment = self.left
            ws.cell(row=current_row, column=3, value=material.get('unit', 'шт.')).alignment = self.center
            ws.cell(row=current_row, column=4, value=material.get('quantity', 1)).alignment = self.center
            ws.cell(row=current_row, column=5, value=material.get('cost_rub', 0)).alignment = self.center
            ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}").alignment = self.center
            current_row += 1
        last_mat_row = current_row - 1
        
        ws.merge_cells(f"B{current_row}:E{current_row}")
        ws[f"B{current_row}"] = "Итого материалы (руб.)"
        ws[f"B{current_row}"].font = self.bold_font
        ws[f"B{current_row}"].alignment = self.left
        ws[f"F{current_row}"] = f"=SUM(F{first_mat_row}:F{last_mat_row})"
        ws[f"F{current_row}"].font = self.bold_font
        ws[f"F{current_row}"].alignment = self.center
        
        self._apply_border(ws, first_mat_row - 1, current_row, 1, 6)
        return current_row
    
    def _write_total_section(self, ws, current_row, total: float, zn_number: str):
        """Итоговая секция с суммой прописью и подписями"""
        current_row += 2
        
        # Сводная таблица
        headers = ["№", "Наименование", "", "", "", "Сумма (руб.)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            if header:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                cell.alignment = self.center
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=1).alignment = self.center
        ws.cell(row=current_row, column=2, value="Работа").alignment = self.left
        ws.cell(row=current_row, column=6, value=total).alignment = self.center
        current_row += 1
        
        ws.merge_cells(f"B{current_row}:E{current_row}")
        ws[f"B{current_row}"] = "Всего к оплате (руб.)"
        ws[f"B{current_row}"].font = self.bold_font
        ws[f"B{current_row}"].alignment = self.left
        ws[f"F{current_row}"] = total
        ws[f"F{current_row}"].font = self.bold_font
        ws[f"F{current_row}"].number_format = '#,##0.00'
        ws[f"F{current_row}"].alignment = self.center
        self._apply_border(ws, current_row - 1, current_row, 1, 6)
        current_row += 1
        
        ws.merge_cells(f"B{current_row}:E{current_row}")
        ws[f"B{current_row}"] = "Всего по заказ-наряду:"
        ws[f"B{current_row}"].font = self.bold_font
        ws[f"B{current_row}"].alignment = self.left
        current_row += 1
        
        amount_words = self._amount_in_words(total)
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws[f"B{current_row}"] = amount_words
        ws[f"B{current_row}"].font = Font(bold=True)
        ws[f"B{current_row}"].alignment = self.left
        current_row += 2
        
        # Подписи
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws[f"B{current_row}"] = "Заказчик________________ МП Исполнитель_______________ МП"
        ws[f"B{current_row}"].alignment = self.center
        current_row += 2
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws[f"B{current_row}"] = "Работы выполнены с использованием запасных частей заказчика"
        ws[f"B{current_row}"].alignment = self.center
        
        return current_row
    
    # ====== ТРИ ЛИСТА ======
    
    def generate_order(self, ws, order: dict):
        """Лист 1: Заказ-наряд (ТОЧНО как в v2.1)"""
        current_row = 1
        
        zn_number = order.get('zn_number', '0')
        excel_zn = f"М-{zn_number}"
        vehicle = order.get('vehicle_plate', '').upper().replace(' ', '')
        date_str = order.get('date', '')
        client_name = order.get('client_name', '')
        
        # === ШАПКА ===
        company = self.contractor.get('company', '').split('ИП ')[-1]
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ    {company}"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"ИНН: {self.contractor.get('inn', '')} ОГРНИП: {self.contractor.get('ogrnip', '')}"
        ws[f'A{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = self.contractor.get('address', '')
        ws[f'A{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"{self.contractor.get('email', '')} {self.contractor.get('phone', '')}"
        ws[f'A{current_row}'].alignment = self.center
        current_row += 1

        current_row += 1
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"ЗАКАЗ – НАРЯД № {excel_zn}"
        ws[f'B{current_row}'].font = Font(bold=True, size=14)
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Дата и время приема заказа: {date_str} г."
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Дата и время окончания работ: {date_str} г."
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1

        # === КЛИЕНТ ===
        client_data = self.clients.get(client_name, {})
        if not client_data:
            client_data = {
                'company': client_name,
                'address': 'Адрес не указан',
                'default_vehicle': 'Автомобиль'
            }

        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = "Заказчик"
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = client_name
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'B{current_row}:F{current_row}')
        client_address = order.get('client_address', '') or client_data.get('address', 'Адрес не указан')
        ws[f'B{current_row}'] = f"Адрес: {client_address}"
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1

        # Марка/модель — из БД
        vehicle_brand = order.get('vehicle_brand', '')
        vehicle_model = order.get('vehicle_model', '')
        vehicle_type = f"{vehicle_brand} {vehicle_model}".strip() or client_data.get('default_vehicle', 'Автомобиль')
        ws[f'B{current_row}'] = f"Марка, модель: {vehicle_type}"
        ws[f'B{current_row}'].alignment = self.left
        ws[f'E{current_row}'] = "Двигатель №"
        ws[f'E{current_row}'].alignment = self.center
        current_row += 1

        license_plate = self._format_plate(vehicle)
        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws[f'B{current_row}'] = f"Государственный рег. номер: {license_plate}"
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].alignment = self.left
        ws[f'E{current_row}'] = "Шасси №"
        ws[f'E{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws[f'B{current_row}'] = "VIN"
        ws[f'B{current_row}'].alignment = self.left
        ws[f'E{current_row}'] = "Кузов №"
        ws[f'E{current_row}'].alignment = self.center
        current_row += 1

        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Выполненные работы по заказ–наряду № {excel_zn}"
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1

        # === РАБОТЫ ===
        works_start_row = current_row
        headers = ["№", "Наименование работ", "Норма времени", "Кол-во", "Стоимость (руб.)", "Сумма (руб.)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = self.center
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        works = order.get('work_items', [])
        works_first_row = current_row
        for i, work in enumerate(works, 1):
            ws.cell(row=current_row, column=1, value=i).alignment = self.center
            ws.cell(row=current_row, column=2, value=work['name']).alignment = self.left
            ws.cell(row=current_row, column=3, value=work['norm_hours']).alignment = self.center
            ws.cell(row=current_row, column=4, value=work.get('quantity', 1)).alignment = self.center
            ws.cell(row=current_row, column=5, value=work['rate_rub']).alignment = self.center
            ws.cell(row=current_row, column=6, value=f"=C{current_row}*D{current_row}*E{current_row}").alignment = self.center
            current_row += 1
        works_last_data_row = current_row - 1

        ws.merge_cells(f"B{current_row}:E{current_row}")
        ws[f"B{current_row}"] = "Итого работы (руб.)"
        ws[f"B{current_row}"].font = Font(bold=True)
        ws[f"B{current_row}"].alignment = self.left
        ws[f"F{current_row}"] = f"=SUM(F{works_first_row}:F{works_last_data_row})"
        ws[f"F{current_row}"].font = Font(bold=True)
        ws[f"F{current_row}"].alignment = self.center
        works_total_row = current_row
        current_row += 2

        # === МАТЕРИАЛЫ ===
        materials = order.get('material_items', [])
        if materials:
            materials_start_row = current_row
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = f"Расходная накладная по заказ–наряду № {excel_zn}"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = self.center
            current_row += 1

            headers = ["№", "Наименование", "Ед.", "Кол-во", "Цена (руб.)", "Сумма (руб.)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = self.center
            ws.row_dimensions[current_row].height = 30
            current_row += 1

            materials_first_row = current_row
            for i, material in enumerate(materials, 1):
                ws.cell(row=current_row, column=1, value=i).alignment = self.center
                ws.cell(row=current_row, column=2, value=material['name']).alignment = self.left
                ws.cell(row=current_row, column=3, value=material.get('unit', 'шт.')).alignment = self.center
                ws.cell(row=current_row, column=4, value=material.get('quantity', 1)).alignment = self.center
                ws.cell(row=current_row, column=5, value=material.get('cost_rub', 0)).alignment = self.center
                ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}").alignment = self.center
                current_row += 1
            materials_last_data_row = current_row - 1

            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "Итого материалы (руб.)"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = self.left
            ws[f"F{current_row}"] = f"=SUM(F{materials_first_row}:F{materials_last_data_row})"
            ws[f"F{current_row}"].font = Font(bold=True)
            ws[f"F{current_row}"].alignment = self.center
            materials_total_row = current_row
            current_row += 2
        else:
            materials_total_row = 0
            current_row += 1

        # === ИТОГИ ===
        totals_start_row = current_row
        headers = ["№", "Наименование", "", "", "", "Сумма (руб.)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            if header:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                cell.alignment = self.center
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        ws.cell(row=current_row, column=1, value=1).alignment = self.center
        ws.cell(row=current_row, column=2, value="Работа").alignment = self.left
        ws.cell(row=current_row, column=6, value=f"=F{works_total_row}").alignment = self.center
        work_summary_row = current_row
        current_row += 1

        ws.cell(row=current_row, column=1, value=2).alignment = self.center
        ws.cell(row=current_row, column=2, value="Материалы").alignment = self.left
        if materials_total_row > 0:
            ws.cell(row=current_row, column=6, value=f"=F{materials_total_row}").alignment = self.center
        else:
            ws.cell(row=current_row, column=6, value=0).alignment = self.center
        materials_summary_row = current_row
        current_row += 1

        ws.merge_cells(f"B{current_row}:E{current_row}")
        ws[f"B{current_row}"] = "Всего к оплате (руб.)"
        ws[f"B{current_row}"].font = Font(bold=True)
        ws[f"B{current_row}"].alignment = self.left
        ws[f"F{current_row}"] = f"=F{work_summary_row}+F{materials_summary_row}"
        ws[f"F{current_row}"].font = Font(bold=True)
        ws[f"F{current_row}"].alignment = self.center
        current_row += 1

        ws.merge_cells(f"B{current_row}:E{current_row}")
        ws[f"B{current_row}"] = "Всего по заказ-наряду:"
        ws[f"B{current_row}"].font = Font(bold=True)
        ws[f"B{current_row}"].alignment = self.left
        current_row += 1

        total_amount = order.get('total_amount', 0)
        amount_words = self._amount_in_words(total_amount)
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws[f"B{current_row}"] = amount_words
        ws[f"B{current_row}"].font = Font(bold=True)
        ws[f"B{current_row}"].alignment = self.left
        current_row += 2

        # === ПОДПИСИ ===
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws[f"B{current_row}"] = "Заказчик________________ МП Исполнитель_______________ МП"
        ws[f"B{current_row}"].alignment = self.center
        current_row += 2
        ws.merge_cells(f"B{current_row}:F{current_row}")
        ws[f"B{current_row}"] = "Работы выполнены с использованием запасных частей заказчика"
        ws[f"B{current_row}"].alignment = self.center

        # === ФОРМАТИРОВАНИЕ ГРАНИЦ ===
        ranges_to_format = [(works_start_row, works_total_row)]
        if materials_total_row > 0:
            ranges_to_format.append((materials_start_row, materials_total_row))
        ranges_to_format.append((totals_start_row, totals_start_row + 3))

        for start, end in ranges_to_format:
            for row in range(start, end + 1):
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.border = self.thin_border
        
        # Числовой формат для колонок E и F
        for row in ws.iter_rows():
            for cell in row:
                if cell.column in [5, 6] and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    def generate_schet(self, ws, order: dict):
        """Лист 2: Счёт на оплату"""
        current_row = 1
        zn = order.get('zn_number', '')
        date = order.get('date', '')
        plate = order.get('vehicle_plate', '')
        client_name = order.get('client_name', '')
        
        current_row = self._write_company_header(ws, current_row)
        current_row += 1
        
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"СЧЁТ № {zn} от {date}"
        ws[f'B{current_row}'].font = Font(bold=True, size=14)
        ws[f'B{current_row}'].alignment = self.center
        current_row += 2
        
        # Банковские реквизиты
        bank = self.contractor.get('bank', '')
        bik = self.contractor.get('bik', '')
        account = self.contractor.get('account', '')
        corr = self.contractor.get('corr_account', '')
        
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Банк: {bank}"
        current_row += 1
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"БИК: {bik}"
        current_row += 1
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Р/с: {account}"
        current_row += 1
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Корр/с: {corr}"
        current_row += 2
        
        # Заказчик
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Заказчик: {client_name}"
        ws[f'B{current_row}'].font = self.bold_font
        current_row += 1
        
        formatted_plate = self._format_plate(plate)
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Автомобиль: {formatted_plate}"
        current_row += 1
        
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Основание: Заказ-наряд № М-{zn} от {date}"
        current_row += 1
        
        # Таблицы
        current_row = self._write_work_table(ws, current_row, order.get('work_items', []))
        current_row = self._write_material_table(ws, current_row, order.get('material_items', []))
        
        total = order.get('total_amount', 0)
        self._write_total_section(ws, current_row, total, zn)
    
    def generate_upd(self, ws, order: dict):
        """Лист 3: УПД (статус 1)"""
        current_row = 1
        zn = order.get('zn_number', '')
        date = order.get('date', '')
        plate = order.get('vehicle_plate', '')
        client_name = order.get('client_name', '')
        client_inn = order.get('client_inn', '')
        
        current_row = self._write_company_header(ws, current_row)
        current_row += 1
        
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = "УНИВЕРСАЛЬНЫЙ ПЕРЕДАТОЧНЫЙ ДОКУМЕНТ"
        ws[f'B{current_row}'].font = Font(bold=True, size=14)
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1
        
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = "Статус: 1 (счёт-фактура и передаточный документ)"
        ws[f'B{current_row}'].font = self.bold_font
        ws[f'B{current_row}'].alignment = self.center
        current_row += 1
        
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Счёт-фактура № {zn} от {date}"
        ws[f'B{current_row}'].alignment = self.center
        current_row += 2
        
        # Продавец
        company = self.contractor.get('company', '').replace('ИП ', '')
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = "Продавец (Исполнитель):"
        ws[f'B{current_row}'].font = self.bold_font
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = f"ИП {company}"
        current_row += 1
        
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = "ИНН:"
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = self.contractor.get('inn', '')
        current_row += 1
        
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = "Адрес:"
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = self.contractor.get('address', '')
        current_row += 2
        
        # Покупатель
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = "Покупатель (Заказчик):"
        ws[f'B{current_row}'].font = self.bold_font
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = client_name
        current_row += 1
        
        if client_inn:
            ws.merge_cells(f'B{current_row}:C{current_row}')
            ws[f'B{current_row}'] = "ИНН покупателя:"
            ws.merge_cells(f'D{current_row}:F{current_row}')
            ws[f'D{current_row}'] = client_inn
            current_row += 1
        current_row += 1
        
        # Основание
        formatted_plate = self._format_plate(plate)
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Основание: Заказ-наряд № М-{zn} от {date}"
        current_row += 1
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"Автомобиль: {formatted_plate}"
        current_row += 1
        
        # Таблицы
        current_row = self._write_work_table(ws, current_row, order.get('work_items', []))
        current_row = self._write_material_table(ws, current_row, order.get('material_items', []))
        
        total = order.get('total_amount', 0)
        
        # Итого с НДС
        current_row += 1
        ws.merge_cells(f'B{current_row}:E{current_row}')
        ws[f'B{current_row}'] = "Всего к оплате (без НДС):"
        ws[f'B{current_row}'].font = self.bold_font
        ws[f'B{current_row}'].alignment = self.right
        ws[f'F{current_row}'] = total
        ws[f'F{current_row}'].font = self.bold_font
        ws[f'F{current_row}'].number_format = '#,##0.00'
        ws[f'F{current_row}'].alignment = self.center
        self._apply_border(ws, current_row, current_row, 1, 6)
        current_row += 1
        
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = "НДС не облагается (УСН)"
        ws[f'B{current_row}'].alignment = self.right
        current_row += 1
        
        self._write_total_section(ws, current_row, total, zn)
    
    def generate(self, order: dict) -> str:
        """Генерирует 3 листа. Возвращает имя файла."""
        plate = order.get('vehicle_plate', 'АА0000').replace(' ', '_')
        date_clean = order.get('date', datetime.now().strftime('%d.%m.%Y')).replace('.', '-')
        zn = order.get('zn_number', '0')
        filename = f"ZN_{zn}_{plate}_{date_clean}.xlsx"
        output_path = self.output_dir / filename
        
        wb = Workbook()
        
        # Лист 1: Заказ-наряд
        ws1 = wb.active
        ws1.title = "Заказ-наряд"
        self.generate_order(ws1, order)
        
        # Лист 2: Счёт
        ws2 = wb.create_sheet("Счёт")
        self.generate_schet(ws2, order)
        
        # Лист 3: УПД
        ws3 = wb.create_sheet("УПД")
        self.generate_upd(ws3, order)
        
        # Настройка колонок для всех листов
        for ws in [ws1, ws2, ws3]:
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
        
        wb.save(output_path)
        return str(output_path)
